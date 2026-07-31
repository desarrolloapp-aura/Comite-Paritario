import base64
import os
import subprocess
import docx
import openpyxl
from multiprocessing import context
from django.core.files.base import ContentFile
from django.core.exceptions import ImproperlyConfigured
from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import get_object_or_404, render, redirect
from django.utils import timezone
import threading

def enviar_correo_async(asunto, mensaje, from_email, recipient_list):
    try:
        send_mail(
            subject=asunto,
            message=mensaje,
            from_email=from_email,
            recipient_list=recipient_list,
            fail_silently=True
        )
    except Exception as e:
        print(f"Error al enviar el correo: {e}")
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from app_reportes.models import InspeccionCabecera, ItemInspeccion, ItemInspeccionEvidencia, ObservacionConducta, ItemObservacionConducta, ReporteHSGI, ItemObsConductaEvidencia, ReporteHSGIEvidencia
from datetime import date, datetime, timedelta
from docxtpl import DocxTemplate, InlineImage
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont



# Create your views here.
def ver_index(request):
    return render(request,'index.html')

def ver_inspecciones(request):
    context = {
        'today': date.today(),
    }
    return render(request, 'form_inspecciones.html', context)

def ver_obsconducta(request):
    context = {
            'today': date.today(),
        }
    return render(request,'form_obsconducta.html', context)

def ver_reportehsgi(request):
    context = {
            'today': date.today(),
        }
    return render(request,'form_reportehsgi.html', context)

def evid_inspecciones(request,token):
    inspeccion = get_object_or_404(InspeccionCabecera,token=token)
    items = ItemInspeccion.objects.filter(inspeccion=inspeccion).order_by('numero_item')
    evidencias = ItemInspeccionEvidencia.objects.filter(item__in=items)
    hay_pendientes = items.filter(cumplimiento='PENDIENTE').exists()
    datos = {
        'today': date.today(),
        'inspeccion': inspeccion,
        'items': items,
        'evidencias': evidencias,
        'hay_pendientes': hay_pendientes,
    }
    return render(request, 'evid_inspecciones.html', datos)

def evid_obsconducta(request,token):
    observacion = get_object_or_404(ObservacionConducta,token=token)
    items = ItemObservacionConducta.objects.filter(observacion_conducta=observacion).order_by('numero_item')
    hay_pendientes = items.filter(cumplimiento='PENDIENTE').exists()
    datos = {
        'obsconducta': observacion,
        'items': items,
        'hay_pendientes':hay_pendientes
    }
    return render(request, 'evid_obsconducta.html', datos)

def evid_reportehsgi(request,token):
    reporte = get_object_or_404(ReporteHSGI, token=token)

    datos = {
        'today': date.today(),
        'hsgi': reporte,}
    return render(request, 'evid_reportehsgi.html', datos)

def menu_listas(request):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return redirect('inicio')

    # 2. Si es superusuario, carga el menú de listas normalmente
    return render(request, "menu_listas.html")


def lista_inspecciones(request):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return redirect('inicio')
    filtro = request.GET.get('filtro', 'todos').lower()

    inspecciones = InspeccionCabecera.objects.all().order_by('-fecha_creacion')

    if filtro == 'auditados':
        inspecciones = inspecciones.filter(fecha_inspeccion__isnull=False, hora_inspeccion__isnull=False)
    elif filtro == 'no_auditados':
        inspecciones = inspecciones.filter(fecha_inspeccion__isnull=True) | inspecciones.filter(hora_inspeccion__isnull=True)

    context = {
        'inspecciones': inspecciones,
        'filtro': filtro,
    }
    return render(request, 'lista_inspecciones.html', context)

def lista_obsconducta(request):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return redirect('inicio')
    filtro = request.GET.get('filtro', 'todos').lower()
    
    obsconductas = ObservacionConducta.objects.all().order_by('-fecha_creacion')

    if filtro == 'exportados':
        obsconductas = obsconductas.filter(fecha_exportado__isnull=False, hora_exportado__isnull=False)
    elif filtro == 'no_exportados':
        obsconductas = obsconductas.filter(fecha_exportado__isnull=True) | obsconductas.filter(hora_exportado__isnull=True)

    context = {
        'obsconductas': obsconductas,
        'filtro': filtro,
    }
    return render(request, 'lista_obsconducta.html', context)

def lista_reportehsgi(request):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return redirect('inicio')
    filtro = request.GET.get('filtro', 'todos').lower()
        
    reportes = ReporteHSGI.objects.all().order_by('-fecha_creacion')

    if filtro == 'cerrados':
        reportes = reportes.filter(estado_cierre='CERRADA')
    elif filtro == 'no_cerrados':
        reportes = reportes.filter(estado_cierre='PENDIENTE')

    context = {
        'reportes': reportes,
        'filtro': filtro,
    }
    return render(request, 'lista_reportehsgi.html', context)

def revisar_inspeccion(request, id):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return redirect('inicio')
    inspeccion = InspeccionCabecera.objects.filter(id=id).first()
    items = ItemInspeccion.objects.filter(inspeccion=inspeccion).order_by('numero_item')
    datos = {
        'today': date.today(),
        'inspeccion': inspeccion,
        'items': items,
    }
    return render(request, 'revisar_inspecciones.html', datos)

def revisar_obsconducta(request, id):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return redirect('inicio')
    obsconducta = ObservacionConducta.objects.filter(id=id).first()
    items = ItemObservacionConducta.objects.filter(observacion_conducta=obsconducta).order_by('numero_item')
    datos = {
        'today': date.today(),
        'obsconducta': obsconducta,
        'items': items,
    }
    return render(request, 'revisar_obsconducta.html', datos)

def revisar_reportehsgi(request, id):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return redirect('inicio')
    hsgi = ReporteHSGI.objects.filter(id=id).first()
    datos = {
        'today': date.today(),
        'hsgi': hsgi,
    }
    return render(request, 'revisar_reportehsgi.html', datos)

def ver_firmar_inspecciones(request):
    inspeccion_id = request.GET.get('inspeccion', '')
    return render(request, 'firmar_inspecciones.html', {'inspeccion_id': inspeccion_id})


def ver_firmar_obsconducta_creador(request):
    obsconducta_id = request.GET.get('obsconducta', '')
    return render(request, 'firmar_obsconducta_creador.html', {'obsconducta_id': obsconducta_id})


def ver_firmar_obsconducta_observado(request):
    obsconducta_id = request.GET.get('obsconducta', '')
    return render(request, 'firmar_obsconducta_observado.html', {'obsconducta_id': obsconducta_id})




#################################################
# ENVÍO DE RESPUESTAS DE FORMULARIOS
#################################################

def enviar_inspeccion(request):
    if request.method != 'POST':
        return render(request, 'form_inspecciones.html', {'today': date.today()})

    tipo_inspeccion = request.POST.get('sel_for', '').strip().capitalize()
    if tipo_inspeccion not in {'Formal', 'Informal'}:
        tipo_inspeccion = 'Formal'

    fecha_inspeccion = None
    fecha_raw = request.POST.get('txtfec', '').strip()
    if fecha_raw:
        fecha_inspeccion = date.fromisoformat(fecha_raw)

    hora_inspeccion = None
    hora_raw = request.POST.get('txthor', '').strip()
    if hora_raw:
        hora_inspeccion = datetime.strptime(hora_raw, '%H:%M').time()

    nueva_inspeccion = InspeccionCabecera(
        tipo_inspeccion=tipo_inspeccion,
        faena=request.POST.get('txtfae', ''),
        lugar=request.POST.get('txtlug', ''),
        #fecha_inspeccion=fecha_inspeccion,
        #hora_inspeccion=hora_inspeccion,
        realizada_por=request.POST.get('txtrea', ''),
        notificado_a=request.POST.get('txtnot', ''),
        correo_usuario_creador=request.POST.get('txtema_usu', ''),
    )
    nueva_inspeccion.save()

    observaciones = request.POST.getlist('observaciones[]')
    grados_riesgo = request.POST.getlist('grado_riesgo[]')
    recomendaciones = request.POST.getlist('recomendaciones[]')
    responsables = request.POST.getlist('responsable[]')
    fechas_cumplimiento = request.POST.getlist('fecha[]')

    items_creados = []

    for index, observacion in enumerate(observaciones, start=1):
        if not observacion.strip():
            continue

        fecha_cumplimiento = None
        if index <= len(fechas_cumplimiento) and fechas_cumplimiento[index - 1]:
            fecha_cumplimiento = date.fromisoformat(fechas_cumplimiento[index - 1])

        item = ItemInspeccion.objects.create(
            inspeccion=nueva_inspeccion,
            numero_item=index,
            observaciones=observacion,
            grado_riesgo=grados_riesgo[index - 1] if index <= len(grados_riesgo) else '',
            recomendaciones=recomendaciones[index - 1] if index <= len(recomendaciones) else '',
            responsable_accion=responsables[index - 1] if index <= len(responsables) else '',
            fecha_cumplimiento=fecha_cumplimiento,
        )
        items_creados.append(item)

        evidencia_files = request.FILES.getlist(f'evidencia_{index}')

        print(evidencia_files)  # Agrega esta línea para depuración
        if evidencia_files:
            for evidencia_file in evidencia_files:
                ItemInspeccionEvidencia.objects.create(
                    item=item, 
                    archivo=evidencia_file
                )
                ItemInspeccion.objects.filter(id=item.id).update(cumplimiento='SI', fecha_cierre=datetime.now())

    correo_destino = nueva_inspeccion.correo_usuario_creador

    if correo_destino:
        url_subir_evidencia = f"https://reportes-sgi.onrender.com/inspecciones/evidencia_inspeccion/{nueva_inspeccion.token}"

        resumen_tabla = ""
        for it in items_creados:
            resumen_tabla += (
                f"- Ítem N°{it.numero_item}: {it.observaciones}\n"
                f"  • Riesgo: {it.grado_riesgo} | Responsable: {it.responsable_accion}\n"
                f"  • Fecha Límite: {it.fecha_cumplimiento}\n\n"
            )
        asunto = f"Respuesta enviada a Inspección - {nueva_inspeccion.faena}"
        
        mensaje = f"""
Estimado/a {nueva_inspeccion.realizada_por},

Se ha registrado exitosamente la inspección con los siguientes detalles:

DATOS GENERALES:
- Tipo: {nueva_inspeccion.tipo_inspeccion}
- Faena: {nueva_inspeccion.faena}
- Lugar: {nueva_inspeccion.lugar}
- Realizada por: {nueva_inspeccion.realizada_por}
- Fecha: {nueva_inspeccion.fecha_inspeccion}

ÍTEMS Y OBSERVACIONES REGISTRADAS:
{resumen_tabla}
Favor ingresar al sistema para adjuntar evidencias y dar cierre a las observaciones:
{url_subir_evidencia}

Atentamente,
Sistema de Gestión Integrado (SGI)
"""

        try:
            # 2. Enviamos el correo en segundo plano
            threading.Thread(
                target=enviar_correo_async,
                args=(asunto, mensaje, settings.DEFAULT_FROM_EMAIL, [correo_destino])
            ).start()
        except Exception as e:
            # Si falla el hilo, la inspección de todas formas queda guardada
            print(f"Error al iniciar hilo de correo: {e}")

    # =========================================================================

    return render(
        request,
        'form_inspecciones.html',
        {'mensaje': 'Inspección enviada con éxito.', 'today': date.today()},
    )

def enviar_obsconducta(request):
    if request.method != 'POST':
        return render(request, 'form_obsconducta.html')

    observacion_planificada = request.POST.get('sel_pla', 'SI').strip().upper()
    if observacion_planificada not in {'SI', 'NO'}:
        observacion_planificada = 'SI'

    mejora = request.POST.get('sel_mej', 'SI').strip().upper()
    if mejora not in {'SI', 'NO'}:
        mejora = 'SI'

    fecha_observacion = None
    fecha_raw = request.POST.get('txtfec', '').strip()
    if fecha_raw:
        fecha_observacion = date.fromisoformat(fecha_raw)

    descripcion_tarea = request.POST.get('txtdes', '').strip()
    observacion_texto = request.POST.get('txtobs', '').strip()


    nueva_obsconducta = ObservacionConducta(
        area_trabajo=request.POST.get('txtare', '').strip(),
        persona_observada=request.POST.get('txtper', '').strip(),
        fecha_observacion=fecha_observacion,
        descripcion_tarea=descripcion_tarea,
        observacion=observacion_texto,
        antiguedad_puesto=request.POST.get('txtant', '').strip(),
        tarea=request.POST.get('txttar', '').strip(),
        observacion_planificada=observacion_planificada,
        mejora=mejora,
        nombre_usuario_creador=request.POST.get('txtnom_usu', '').strip(),
        correo_usuario_creador=request.POST.get('txtema_usu', '').strip(),
    )
    nueva_obsconducta.save()

    analisis_causas = request.POST.getlist('analisis[]')
    acciones_correctivas = request.POST.getlist('acciones[]')
    responsables = request.POST.getlist('responsable[]')
    fechas_cumplimiento = request.POST.getlist('fecha[]')

    items_creados = []
    for index, analisis in enumerate(analisis_causas, start=1):
        if not analisis.strip() and not acciones_correctivas[index - 1].strip() and not responsables[index - 1].strip():
            continue

        fecha_cumplimiento = None
        if index <= len(fechas_cumplimiento) and fechas_cumplimiento[index - 1]:
            fecha_cumplimiento = date.fromisoformat(fechas_cumplimiento[index - 1])
        if not fecha_cumplimiento:
            fecha_cumplimiento = date.today()

        item = ItemObservacionConducta.objects.create(
            observacion_conducta=nueva_obsconducta,
            numero_item=index,
            analisis_causa=analisis,
            acciones_correctivas=acciones_correctivas[index - 1] if index <= len(acciones_correctivas) else '',
            responsable=responsables[index - 1] if index <= len(responsables) else '',
            fecha_cumplimiento=fecha_cumplimiento,
        )
        items_creados.append(item)

        evidencia_files = request.FILES.getlist(f'evidencia_{index}')
        
        print(evidencia_files)  # Agrega esta línea para depuración
        if evidencia_files:
            for evidencia_file in evidencia_files:
                ItemObsConductaEvidencia.objects.create(
                    item=item, 
                    archivo=evidencia_file
                )

    correo_destino = nueva_obsconducta.correo_usuario_creador

    if correo_destino:
        url_subir_evidencia = f"https://reportes-sgi.onrender.com/obsconducta/evidencia_obsconducta/{nueva_obsconducta.token}"

        resumen_tabla = ""
        for it in items_creados:
            resumen_tabla += (
                f"- Ítem N°{it.numero_item}:\n"
                f"  • Análisis de causa: {it.analisis_causa} \n"
                f"  • Acciones Correctivas: {it.acciones_correctivas} \n"
                f"  • Responsable: {it.responsable} \n"
                f"  • Fecha Límite: {it.fecha_cumplimiento}\n\n"
            )
        asunto = f"Respuesta enviada a Observación de Conducta - {nueva_obsconducta.area_trabajo}"
        
        mensaje = f"""
Estimado/a {nueva_obsconducta.nombre_usuario_creador},

Se ha registrado exitosamente la observación de conducta con los siguientes detalles:

DATOS GENERALES:
- Persona observada: {nueva_obsconducta.persona_observada}
- Antigüedad: {nueva_obsconducta.antiguedad_puesto}
- Área de trabajo: {nueva_obsconducta.area_trabajo}
- Tarea: {nueva_obsconducta.tarea}
- Fecha: {nueva_obsconducta.fecha_observacion}
- Observación planificada: {nueva_obsconducta.observacion_planificada}

ÍTEMS Y OBSERVACIONES REGISTRADAS:
{resumen_tabla}
Favor ingresar al sistema para firmar y dar cierre a las observaciones:
{url_subir_evidencia}

Atentamente,
Sistema de Gestión Integrado (SGI)
"""

        try:
            # 2. Enviamos el correo
            send_mail(
                subject=asunto,
                message=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[correo_destino],
                fail_silently=True  # Pon en False si deseas ver errores de SMTP en consola durante pruebas
            )
        except Exception as e:
            # Si falla el servidor de correo, la inspección de todas formas queda guardada
            print(f"Error al enviar el correo: {e}")
 
    # =========================================================================

    return render(
        request,
        'form_obsconducta.html',
        {'mensaje': 'Observación de Conducta enviada con éxito.'},
    )

def enviar_reportehsgi(request):
    if request.method != 'POST':
        return render(request, 'form_reportehsgi.html', {'today': date.today()})

    ambito = request.POST.getlist('selamb', '')


    causa = request.POST.getlist('selcau', '')


    nivel = request.POST.get('selniv', '')
    if nivel not in {'Alto', 'Medio', 'Oportunidad', 'Positivo'}:
        nivel = 'Alto'

    fecha_hal = None
    fecha_raw = request.POST.get('txtfec', '').strip()
    if fecha_raw:
        fecha_hal = date.fromisoformat(fecha_raw)

    hora_hal = None
    hora_raw = request.POST.get('txthor', '').strip()
    if hora_raw:
        hora_hal = datetime.strptime(hora_raw, '%H:%M').time()

    fecha_cie = None
    fecha_raw = request.POST.get('txtfec_cierre', '').strip()
    if fecha_raw:
        fecha_cie = date.fromisoformat(fecha_raw)

    evidencia_files = request.FILES.getlist(f'evidencia')
    
    

    nueva_hsgi = ReporteHSGI(
        cargo=request.POST.get('txtcar', ''),
        centro_trabajo=request.POST.get('txtcen', ''),
        fecha_hallazgo=fecha_hal,
        hora_hallazgo=hora_hal,
        turno=request.POST.get('txttur', ''),
        proceso_hallazgo=request.POST.get('txtpro', ''),
        actividad_involucrada=request.POST.get('txtact', ''),
        lugar_especifico=request.POST.get('txtlug', ''),
        ambito_hallazgo=ambito,
        causa=causa,
        nivel_hallazgo=nivel,
        supervisor_hallazgo=request.POST.get('txtsup', ''),
        fecha_cierre=fecha_cie,
        descripcion_hallazgo=request.POST.get('txtdes', ''),
        accion_inmediata=request.POST.get('txtacc', ''),
        responsable_cierre=request.POST.get('txtresp', ''),
        nombre_usuario_creador=request.POST.get('txtnom_usu', ''),
        correo_usuario_creador=request.POST.get('txtema_usu', ''),
    )
    nueva_hsgi.save()

    if evidencia_files:
        # 1. Guardar cada imagen recibida
        for evidencia_file in evidencia_files:
            ReporteHSGIEvidencia.objects.create(
                reporte_hsgi=nueva_hsgi, archivo=evidencia_file
            )

        # 2. Solo si se guardaron evidencias, marcamos como cerrado
        nueva_hsgi.estado_cierre = 'CERRADA'
        nueva_hsgi.save()

    correo_destino = nueva_hsgi.correo_usuario_creador

    if correo_destino:
        url_subir_evidencia = f"https://reportes-sgi.onrender.com/reportehsgi/evidencia_hsgi/{nueva_hsgi.token}"


        asunto = f"Respuesta enviada a Reportes HSGI - {nueva_hsgi.centro_trabajo}"
        
        mensaje = f"""
Estimado/a {nueva_hsgi.nombre_usuario_creador},

Se ha registrado exitosamente el reporte HSGI con los siguientes detalles:

DATOS GENERALES:
- Cargo: {nueva_hsgi.cargo}
- Centro de trabajo: {nueva_hsgi.centro_trabajo}
- Fecha hallazgo: {nueva_hsgi.fecha_hallazgo}
- Hora hallazgo: {nueva_hsgi.hora_hallazgo}
- Nivel hallazgo: {nueva_hsgi.nivel_hallazgo}
- Ámbito hallazgo: {nueva_hsgi.ambito_hallazgo}
- Causa: {nueva_hsgi.causa}


Favor ingresar al sistema para adjuntar evidencias y dar cierre a las observaciones:
{url_subir_evidencia}

Atentamente,
Sistema de Gestión Integrado (SGI)
"""

        try:
            # 2. Enviamos el correo
            send_mail(
                subject=asunto,
                message=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[correo_destino],
                fail_silently=True  # Pon en False si deseas ver errores de SMTP en consola durante pruebas
            )
        except Exception as e:
            # Si falla el servidor de correo, la inspección de todas formas queda guardada
            print(f"Error al enviar el correo: {e}")

    # =========================================================================

    return render(
        request,
        'form_reportehsgi.html',
        {'mensaje': 'Reporte HSGI enviado con éxito.', 'today': date.today()},
    )
    
def enviar_inspeccion_evidencia(request, token, item_id):
    item = get_object_or_404(ItemInspeccion, id=item_id)
    if request.method == 'POST':
        

        # Captura las evidencias enviadas con la clave evidencia_N
        evidencia_files = request.FILES.getlist(f'evidencia_{item.numero_item}')

        if evidencia_files:
            # 1. Guardar cada imagen recibida
            for evidencia_file in evidencia_files:
                ItemInspeccionEvidencia.objects.create(
                    item=item, archivo=evidencia_file
                )

            # 2. Solo si se guardaron evidencias, marcamos como cerrado
            item.cumplimiento = 'SI'
            item.fecha_cierre = datetime.now()
            item.save()

        return redirect(
            reverse('evidencia_inspecciones', args=[token])
        )

    return redirect(reverse('evidencia_inspecciones', args=[token]))


def enviar_obsconducta_evidencia(request, token, item_id):
    item = get_object_or_404(ItemObservacionConducta, id=item_id)
    if request.method == 'POST':
        

        # Captura las evidencias enviadas con la clave evidencia_N
        evidencia_files = request.FILES.getlist(f'evidencia_{item.numero_item}')

        if evidencia_files:
            # 1. Guardar cada imagen recibida
            for evidencia_file in evidencia_files:
                ItemObsConductaEvidencia.objects.create(
                    item=item, archivo=evidencia_file
                )

            # 2. Solo si se guardaron evidencias, marcamos como cerrado
            item.cumplimiento = 'SI'
            item.save()

        return redirect(
            reverse('evidencia_obsconducta', args=[token])
        )

    return redirect(reverse('evidencia_obsconducta', args=[token]))

def enviar_reportehsgi_evidencia(request, token):
    hsgi = get_object_or_404(ReporteHSGI, token=token)
    if request.method == 'POST':
        
    
        evidencia_files = request.FILES.getlist(f'evidencia')

        if evidencia_files:
            # 1. Guardar cada imagen recibida
            for evidencia_file in evidencia_files:
                ReporteHSGIEvidencia.objects.create(
                    reporte_hsgi=hsgi, archivo=evidencia_file
                )

            # 2. Solo si se guardaron evidencias, marcamos como cerrado
            hsgi.estado_cierre = 'CERRADA'
            hsgi.save()

        return redirect(
            reverse('evidencia_reportehsgi', args=[token])
        )
    
    return redirect(reverse('evidencia_reportehsgi', args=[token]))

def guardar_firma_inspeccion(request, token):
    if request.method == 'POST':
        firma_data = request.POST.get('firma_data', '')
        inspeccion = get_object_or_404(InspeccionCabecera, token=token)
        if inspeccion and firma_data:
            header, encoded = firma_data.split(',', 1)
            format_name = header.split(';')[0].split(':')[1]
            extension = format_name.split('/')[-1]
            nombre_archivo = f'firma_{inspeccion.realizada_por}_inspec_{inspeccion.id}.{extension}'
            firma_bytes = base64.b64decode(encoded)
            inspeccion.firma_responsable.save(nombre_archivo, ContentFile(firma_bytes), save=True)
            return JsonResponse({'ok': True, 'archivo': inspeccion.firma_responsable.url})
    return JsonResponse({'ok': False}, status=400)

def guardar_firma_obsconducta_creador(request, token):
    if request.method == 'POST':
        firma_data = request.POST.get('firma_data', '')
        obsconducta = get_object_or_404(ObservacionConducta,token=token)
        if obsconducta and firma_data:
            header, encoded = firma_data.split(',', 1)
            format_name = header.split(';')[0].split(':')[1]
            extension = format_name.split('/')[-1]
            nombre_archivo = f'firma_creador_{obsconducta.nombre_usuario_creador}_obscon_{obsconducta.id}.{extension}'
            firma_bytes = base64.b64decode(encoded)
            obsconducta.firma_creador.save(nombre_archivo, ContentFile(firma_bytes), save=True)
            return JsonResponse({'ok': True, 'archivo': obsconducta.firma_creador.url})
    return JsonResponse({'ok': False}, status=400)

def guardar_firma_obsconducta_observado(request, token):
    if request.method == 'POST':
        firma_data = request.POST.get('firma_data', '')
        obsconducta = get_object_or_404(ObservacionConducta,token=token)
        if obsconducta and firma_data:
            header, encoded = firma_data.split(',', 1)
            format_name = header.split(';')[0].split(':')[1]
            extension = format_name.split('/')[-1]
            nombre_archivo = f'firma_observado_{obsconducta.persona_observada}_obscon_{obsconducta.id}.{extension}'
            firma_bytes = base64.b64decode(encoded)
            obsconducta.firma_observado.save(nombre_archivo, ContentFile(firma_bytes), save=True)
            return JsonResponse({'ok': True, 'archivo': obsconducta.firma_observado.url})
    return JsonResponse({'ok': False}, status=400)


def exportar_inspeccion_pdf(request, inspeccion_id):
    try:
        inspeccion = InspeccionCabecera.objects.get(id=inspeccion_id)
        fecha_raw = datetime.now().date().isoformat()  # Valor por defecto si no se proporciona
        if fecha_raw:
            
            fecha_format = date.fromisoformat(fecha_raw)

        hora_raw = datetime.now().time().strftime('%H:%M')  # Valor por defecto si no se proporciona
        if hora_raw:
            
            hora_format = datetime.strptime(hora_raw, '%H:%M').time()

        
        inspeccion.fecha_inspeccion = fecha_format
        inspeccion.hora_inspeccion = hora_format
        inspeccion.save()
        items = ItemInspeccion.objects.filter(inspeccion_id=inspeccion).order_by('numero_item')
        iteminspeccion = []
        for item in items:
            iteminspeccion.append(            {
                    'numero_item': item.numero_item,
                    'observaciones': item.observaciones,
                    'grado_riesgo': item.grado_riesgo,
                    'recomendaciones': item.recomendaciones,
                    'responsable_accion': item.responsable_accion,
                    'fecha_cumplimiento': item.fecha_cumplimiento.strftime('%d/%m/%Y') if item.fecha_cumplimiento else '',
                    'cumplimiento': item.cumplimiento,
                })


        # 1. Cargar la plantilla Word
        ruta=os.path.join(settings.BASE_DIR,"app_reportes", "static", "plantillas", "plantillaInspecciones.docx")
        doc = DocxTemplate(ruta)
        ruta_firma = os.path.join(settings.BASE_DIR, 'media', inspeccion.firma_responsable.name) if inspeccion.firma_responsable else None
        firma_responsable = InlineImage(doc, ruta_firma, width=docx.shared.Inches(2)) if ruta_firma else None
        # 2. Armar el diccionario de datos
        contexto = {
            'tipo_inspeccion': inspeccion.tipo_inspeccion,
            'faena': inspeccion.faena,
            'lugar': inspeccion.lugar,
            'fecha_inspeccion': inspeccion.fecha_inspeccion.strftime('%d/%m/%Y'),
            'hora_inspeccion': inspeccion.hora_inspeccion.strftime('%H:%M'),
            'realizada_por': inspeccion.realizada_por,
            'notificado_a': inspeccion.notificado_a,
            'items': iteminspeccion,
            'firma_responsable': firma_responsable,
        }
        
        # 3. Renderizar el Word en memoria
        doc.render(contexto)
        
        ruta_word_temp = f"temp_inspeccion_{inspeccion.id}.docx"
        doc.save(ruta_word_temp)

        docs = os.path.join(settings.BASE_DIR, "media", "docs")
        ruta_pdf_generado = os.path.join(
            docs, f"temp_i_pdf_{inspeccion.id}.pdf"
        )
        try:
            subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                ruta_word_temp, '--outdir', docs
            ], check=True)
            expected_output = os.path.join(docs, f"temp_inspeccion_{inspeccion.id}.pdf")
            if os.path.exists(expected_output):
                os.rename(expected_output, ruta_pdf_generado)
        except Exception as exc:
            if os.path.exists(ruta_word_temp):
                os.remove(ruta_word_temp)
            raise RuntimeError(
                'No se pudo convertir el documento Word a PDF. Verifique que LibreOffice esté instalado y que la plantilla sea compatible.'
            ) from exc
    finally:
        pass
        # Limpieza de archivos temporales
    if os.path.exists(ruta_word_temp):
        os.remove(ruta_word_temp)

    with open(ruta_pdf_generado, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/pdf')
        # 'inline' lo abre en pantalla, 'attachment' lo descarga
        response['Content-Disposition'] = f'inline; filename="Inspeccion_{inspeccion.id}_{datetime.now().strftime("%Y-%m-%d")}.pdf"'
        return response
    
    if os.path.exists(ruta_pdf_generado):
        os.remove(ruta_pdf_generado)


def exportar_obsconducta_pdf(request, obsconducta_id):
    try:
        obsconducta = ObservacionConducta.objects.get(id=obsconducta_id)        
        fecha_raw = datetime.now().date().isoformat()  # Valor por defecto si no se proporciona
        if fecha_raw:
            
            fecha_format = date.fromisoformat(fecha_raw)

        hora_raw = datetime.now().time().strftime('%H:%M')  # Valor por defecto si no se proporciona
        if hora_raw:
            
            hora_format = datetime.strptime(hora_raw, '%H:%M').time()

        
        obsconducta.fecha_exportado = fecha_format
        obsconducta.hora_exportado = hora_format
        obsconducta.save()

        items = ItemObservacionConducta.objects.filter(observacion_conducta_id=obsconducta.id).order_by('numero_item')
        itemobs = []
        for item in items:
            itemobs.append(            {
                    'numero_item': item.numero_item,
                    'analisis': item.analisis_causa,
                    'acciones': item.acciones_correctivas,
                    'responsable': item.responsable,
                    'fecha': item.fecha_cumplimiento.strftime('%d/%m/%Y') if item.fecha_cumplimiento else '',
                })


        # 1. Cargar la plantilla Word
        ruta=os.path.join(settings.BASE_DIR,"app_reportes", "static", "plantillas", "plantillaObservacion-Conducta.docx")
        doc = DocxTemplate(ruta)
        ruta_firma_creador = os.path.join(settings.BASE_DIR, 'media', obsconducta.firma_creador.name) if obsconducta.firma_creador else None
        firma_creador = InlineImage(doc, ruta_firma_creador, width=docx.shared.Inches(2)) if ruta_firma_creador else None
        ruta_firma_observado = os.path.join(settings.BASE_DIR, 'media', obsconducta.firma_observado.name) if obsconducta.firma_observado else None
        firma_observado = InlineImage(doc, ruta_firma_observado, width=docx.shared.Inches(2)) if ruta_firma_observado else None
        # 2. Armar el diccionario de datos
        contexto = {
            'area_trabajo': obsconducta.area_trabajo,
            'antiguedad': obsconducta.antiguedad_puesto,
            'persona': obsconducta.persona_observada,
            'observacion': obsconducta.observacion_planificada,
            'fecha_observacion': obsconducta.fecha_observacion.strftime('%d/%m/%Y'),
            'tarea': obsconducta.tarea,
            'descripcion': obsconducta.descripcion_tarea,
            'observacion_texto': obsconducta.observacion,
            'items': itemobs,
            'mejora': obsconducta.mejora,
            'firma_observador': firma_creador,
            'firma_observado': firma_observado,
        }
        
        # 3. Renderizar el Word en memoria
        doc.render(contexto)
        
        ruta_word_temp = f"temp_obsconducta_{obsconducta.id}.docx"
        doc.save(ruta_word_temp)
        docs = os.path.join(settings.BASE_DIR, "media", "docs")
        ruta_pdf_generado = os.path.join(
            docs, f"temp_odc_pdf_{obsconducta.id}.pdf"
        )
        try:
            subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                ruta_word_temp, '--outdir', docs
            ], check=True)
            expected_output = os.path.join(docs, f"temp_obsconducta_{obsconducta.id}.pdf")
            if os.path.exists(expected_output):
                os.rename(expected_output, ruta_pdf_generado)
        except Exception as exc:
            if os.path.exists(ruta_word_temp):
                os.remove(ruta_word_temp)
            raise RuntimeError(
                'No se pudo convertir el documento Word a PDF. Verifique que LibreOffice esté instalado y que la plantilla sea compatible.'
            ) from exc
    finally:
        pass
        # Limpieza de archivos temporales
    if os.path.exists(ruta_word_temp):
        os.remove(ruta_word_temp)

    with open(ruta_pdf_generado, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/pdf')
        # 'inline' lo abre en pantalla, 'attachment' lo descarga
        response['Content-Disposition'] = f'inline; filename="Observacion-Conducta_{obsconducta.id}_{datetime.now().strftime("%Y-%m-%d")}.pdf"'
        return response
    
    if os.path.exists(ruta_pdf_generado):
        os.remove(ruta_pdf_generado)

        

def exportar_reportehsgi_pdf(request, hsgi_id):
    try:
        hsgi = ReporteHSGI.objects.get(id=hsgi_id)
        fecha_raw = datetime.now().date().isoformat()  # Valor por defecto si no se proporciona
        if fecha_raw:
            
            fecha_format = date.fromisoformat(fecha_raw)

        hora_raw = datetime.now().time().strftime('%H:%M')  # Valor por defecto si no se proporciona
        if hora_raw:
            
            hora_format = datetime.strptime(hora_raw, '%H:%M').time()
        hsgi.fecha_exportado = fecha_format
        hsgi.hora_exportado = hora_format
        hsgi.save()
        ruta=os.path.join(settings.BASE_DIR,"app_reportes", "static", "plantillas", "plantillaReporte-HSGI.xlsx")

        wb = openpyxl.load_workbook(ruta)
        ws = wb.active 


        #logo = Image('app_reportes/static/imagenes/aura_ingeniera_s_a_logo.jpg')
   

        # Redimensionar manteniendo la proporción original
        #proporcion = min(150 / logo.width, 60 / logo.height)
        #logo.width = int(logo.width * proporcion)
        #logo.height = int(logo.height * proporcion)
        
        #ws.add_image(logo, 'C3')

        ws['F8'] = hsgi.centro_trabajo

        ws['F9'] = hsgi.nombre_usuario_creador
        ws['S9'] = hsgi.cargo

        ws['E10'] = hsgi.fecha_hallazgo.strftime('%d/%m/%Y') if hsgi.fecha_hallazgo else ''

        normal = InlineFont(u="none")
        subrayado = InlineFont(u="single",b=True)
        turno_rich = CellRichText()
        turno_rich.append(TextBlock(normal,"Turno: "))
        turno_rich.append(TextBlock(subrayado,hsgi.turno))
        ws['H10'] = turno_rich
        hora_rich = CellRichText()
        hora_rich.append(TextBlock(normal,"Hora: "))
        hora_rich.append(TextBlock(subrayado,hsgi.hora_hallazgo.strftime('%H:%M')))
        ws['K10'] = hora_rich
        lugar_rich = CellRichText()
        lugar_rich.append(TextBlock(normal,"Lugar Específico: "))
        lugar_rich.append(TextBlock(subrayado,hsgi.lugar_especifico))
        ws['N10'] = lugar_rich

        proceso_rich = CellRichText()
        proceso_rich.append(TextBlock(normal,"Proceso en que se detecta Hallazgo: "))
        proceso_rich.append(TextBlock(subrayado,hsgi.proceso_hallazgo))
        ws['C11'] = proceso_rich

        actividad_rich = CellRichText()
        actividad_rich.append(TextBlock(normal,"Actividad en que se detecta Hallazgo: "))
        actividad_rich.append(TextBlock(subrayado,hsgi.actividad_involucrada))
        ws['C12'] = actividad_rich

        ambito = str(hsgi.ambito_hallazgo or [])
        causa = str(hsgi.causa or [])

        nivel = hsgi.nivel_hallazgo

        estado = hsgi.estado_cierre

        ws['D16'] = 'X' if 'Seguridad' in ambito else ''
        ws['F16'] = 'X' if 'Salud' in ambito else ''
        ws['I16'] = 'X' if 'Ambiente' in ambito else ''
        ws['K16'] = 'X' if 'Calidad' in ambito else ''

        ws['N16'] = 'X' if 'Acción' in causa else ''
        ws['Q16'] = 'X' if 'Condición' in causa else ''

        ws['V14'] = 'X' if nivel == 'Alto' else ''
        ws['V15'] = 'X' if nivel == 'Medio' else ''
        ws['V16'] = 'X' if nivel == 'Oportunidad' else ''
        ws['V17'] = 'X' if nivel == 'Positivo' else ''

        supervisor_rich = CellRichText()
        supervisor_rich.append(TextBlock(normal,"Supervisor a quien reporta el hallazgo: "))
        supervisor_rich.append(TextBlock(subrayado,hsgi.supervisor_hallazgo))
        ws['C19'] = supervisor_rich
        fecha_rich = CellRichText()
        fecha_rich.append(TextBlock(normal,"Fecha cierre hallazgo: "))
        fecha_rich.append(TextBlock(subrayado,hsgi.fecha_cierre.strftime('%d/%m/%Y')))
        ws['R19'] = fecha_rich

        ws['C22'] = hsgi.descripcion_hallazgo

        ws['C35'] = hsgi.accion_inmediata

        ws['F40'] = hsgi.responsable_cierre

        ws['T40'] = 'X' if estado == 'CERRADA' else ''
        ws['V40'] = 'X' if estado == 'PENDIENTE' else ''

        ruta_temp_excel = os.path.abspath(f'temp_hsgi_{hsgi.id}.xlsx')
        wb.save(ruta_temp_excel)

        
        docs = os.path.join(settings.BASE_DIR, "media", "docs")
        ruta_pdf_generado = os.path.join(docs, f"temp_hsgi_{hsgi.id}.pdf")
        
        try:
            subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                ruta_temp_excel, '--outdir', docs
            ], check=True)
            expected_output = os.path.join(docs, f"temp_hsgi_{hsgi.id}.pdf")
            if os.path.exists(expected_output):
                os.rename(expected_output, ruta_pdf_generado)
        except Exception as exc:
            if os.path.exists(ruta_temp_excel):
                os.remove(ruta_temp_excel)
            raise RuntimeError(
                'No se pudo convertir el documento Excel a PDF. Verifique que LibreOffice esté instalado y que la plantilla sea compatible.'
            ) from exc
    finally:
        pass
        # Limpieza de archivos temporales
    if os.path.exists(ruta_temp_excel):
        os.remove(ruta_temp_excel)

    with open(ruta_pdf_generado, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/pdf')
        # 'inline' lo abre en pantalla, 'attachment' lo descarga
        response['Content-Disposition'] = f'inline; filename="Reporte_HSGI_{hsgi.id}_{datetime.now().strftime("%Y-%m-%d")}.pdf"'
        return response
    
    if os.path.exists(ruta_pdf_generado):
        os.remove(ruta_pdf_generado)
        